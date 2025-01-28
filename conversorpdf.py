from ttkbootstrap import ttk,Window
from ttkbootstrap.constants import *
from tkinter import filedialog
from tkinter import PhotoImage
from docx2pdf import convert
import os
import sys
from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import threading
from time import sleep 

class Aplication: 
    
    def __init__(self, master=None):
        
        self.widget1 = ttk.Frame(master)
        self.widget1.pack(fill=BOTH, expand=True, padx=10, pady=10)
        
        self.lock = True
        self.thread = None
        
        self.msg = ttk.Label(self.widget1, text="Escolha o arquivo", font=("Arial", 11, "bold"))
        self.msg.pack(pady=10)
        
        self.caminhoArquivo = None
        self.caminhoResultado = None
        
        self.arquivo = ttk.Button(self.widget1, text="Procurar", bootstyle=INFO, command=self.procurarArquivo)
        self.arquivo.pack(pady=5)
        
        self.msgArquivo = ttk.Label(self.widget1, text="", font=("Arial", 8, "bold"))
        self.msgArquivo.pack(pady=5)
        
        self.converter_pdf = ttk.Button(self.widget1, text="Converter Word", bootstyle=(INFO, OUTLINE), command=self.converterArquivoPdf,padding = (6,6))
        self.converter_pdf.pack(pady=40)
        
        self.converter_xlsx = ttk.Button(self.widget1, text="Converter Exel", bootstyle=(SUCCESS, OUTLINE), command = self.converterArquivoExel,padding = (9,6))
        self.converter_xlsx.pack(pady=5)
        
  
        self.aviso = ttk.Label(self.widget1, text="", font=("Arial", 11, "bold"))
        self.aviso.pack(pady=10)
        
        #colocar imagem ta dando problema na hora de rodar o executavel (permission danied)
        '''
        caminho_da_imagem = resource_path("pdf.png")
        
        self.imagemG = PhotoImage(file=caminho_da_imagem) 
        self.imagem = self.imagemG.subsample(4, 4)
        self.image_label = ttk.Label(self.widget1, image=self.imagem)
        self.image_label.place(x=300, y=190)
        '''
    
    
    def doc_para_pdf(self, docx_path, pdf_path):
        
        pdf_path_nome = nome_exel(docx_path)
        
        
        try: 
                 
            convert(docx_path, pdf_path)
            print(f'Arquivo convertido para PDF: {pdf_path}')
            self.aviso.config(text="Arquivo convertido!", foreground="green")
            self.caminhoArquivo = None
            self.caminhoResultado = None
            self.msgArquivo["text"] = ""
            
        except Exception as e:
            print(e)
            self.aviso.config(text=f"Erro ao converter o arquivo {pdf_path_nome}", foreground="red")
            self.msgArquivo["text"] = ""
          
            
    def exel_para_pdf(self, exel_path, pdf_path):
        
        pdf_path_nome = nome_exel(exel_path)
        
        pdf_path = pdf_path + '/' + pdf_path_nome.replace('.xlsx','') + '.pdf'
        #pdf_path = pdf_path + 'convertido.pdf'
        
        try:    
            workbook = load_workbook(exel_path)
            sheet = workbook.active

            # Configura o PDF
            pdf = canvas.Canvas(pdf_path, pagesize=A4)
            width, height = A4
            margin = 40  # Margem na página

            y_position = height - margin  # Posição inicial no topo da página

            # Itera sobre as células e escreve no PDF
            for row in sheet.iter_rows(values_only=True):
                text = " | ".join(str(cell) if cell is not None else "" for cell in row)

                # Verifica se a posição vertical ainda cabe na página
                if y_position < margin:
                    pdf.showPage()
                    y_position = height - margin

                pdf.drawString(margin, y_position, text)
                y_position -= 20  # Move para a próxima linha

            pdf.save()
            print(f'Arquivo convertido para PDF: {pdf_path}')
            self.aviso.config(text="Arquivo convertido!", foreground="green")
            self.caminhoArquivo = None
            self.caminhoResultado = None
            self.msgArquivo["text"] = ""
        
        except Exception as e:
            
            print(e)
            self.aviso.config(text=f"Erro ao converter o arquivo {pdf_path_nome}", foreground="red")
            self.msgArquivo["text"] = ""
            
            
    def procurarArquivo(self):
        if self.aviso["text"]:
            self.aviso["text"] = ""
        
        caminho_arquivo = filedialog.askopenfilename(title="Selecione um arquivo")
        
        if caminho_arquivo:
            self.caminhoArquivo = caminho_arquivo
            self.msgArquivo["text"] = self.caminhoArquivo
    
    def validacao(self):
        
        if self.caminhoArquivo is None:
            self.aviso.config(text="Nenhum arquivo selecionado!", foreground ="red")
            return
        
        if self.aviso["text"]:
            self.aviso["text"] = ""
        
        caminho_resultado = filedialog.askdirectory(title="Selecione um diretório")
        
        return caminho_resultado
        
        
    def converterArquivoPdf(self):
        
        caminho_resultado = self.validacao()
        
        if caminho_resultado:
            self.caminhoResultado = caminho_resultado
            self.doc_para_pdf(self.caminhoArquivo, self.caminhoResultado)
            
    def converterArquivoExel(self):
        
        caminho_resultado = self.validacao()
        
        if caminho_resultado:
            self.caminhoResultado = caminho_resultado
            self.exel_para_pdf(self.caminhoArquivo,self.caminhoResultado)
            

def nome_exel(caminho):
        
        nome = [] 
        for i in caminho[::-1]:
            if i == '/':
                break
        
            nome.append(i)    

        resultado = ''.join(nome[::-1])
        return resultado

# Função para obter o caminho correto do arquivo
def resource_path(relative_path):
    if hasattr(sys, '_MEIPASS'):
        # Caminho do diretório temporário do PyInstaller
        return os.path.join(sys._MEIPASS, relative_path)
    return os.path.join(os.path.abspath("."), relative_path)

# Use resource_path para carregar o arquivo




root = Window(themename="darkly")
root.title("Conversor para PDF @ eugabrielbr")
root.resizable(False, False)
root.geometry("400x350")
Aplication(root)
root.mainloop()
